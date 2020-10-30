from selenium import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl as px


#--------------[xlsx設定/h]---------------------
book = px.Workbook()
book.save('sample.xlsx')
wb = px.Workbook()
ws = wb.active
ws.title = "Sheet_1"
#--------------[xlsx設定/r]---------------------

#--------------[ログイン/h]--------------------------
def login(myId, myPass, input_index):
  lgId = driver.find_element_by_id("loginId")
  lgId.send_keys(myId)
  lgpass = driver.find_element_by_id('password')
  lgpass.send_keys(myPass)
  lgbtn = driver.find_element_by_id('loginBtn')
  lgbtn.click()
  Alert(driver).accept()
#--------------[ログイン/r]--------------------------

#--------------[業種サーチ/h]--------------------------
def search(x):
  element = driver.find_element_by_id('srchWord')
  element.send_keys(x)
  srbtn = driver.find_element_by_id('srchButtonUp')
  srbtn.click()
  time.sleep(2)
#--------------[業種サーチ/r]--------------------------

#-------------[繰り返し処理/h]---------------------------------
def fac_info():

  # ループ開始
  for num in range(int(input_index)):
    num+=1
    ws["A"+str(num)] = str(num)
    #企業名-----------------------------------------------------------
    try:
      name = driver.find_element_by_tag_name('h1').text
    except NoSuchElementException:
      name = "None"
    ws["B"+str(num)] = str(name)
    #本社所在地----------------------------------------------------------
    try:
      address = driver.find_element_by_id('corpDescDtoListDescText50').text
    except NoSuchElementException:
      address = "None"
    ws["C"+str(num)] = str(address)
    #売り上げ-----------------------------------------------------------
    try:
      sales = driver.find_element_by_id('corpDescDtoListDescText300').text
    except NoSuchElementException:
      sales = "None"
    ws["D"+str(num)] = str(sales)
    #URL-----------------------------------------------------------
    try:
      uRl = driver.find_element_by_id('corpDescDtoListDescText120').text
    except NoSuchElementException:
      uRl = "None"
    ws["E"+str(num)] = str(uRl)

    # nextページへのID取得
    next_factory = driver.find_element_by_id('next')
    try:
      next_factory.click()
    except NoSuchElementException:
      print("上限到達："+str(num)+"企業をサーチしました")
    wb.save('sample2.xlsx')
    #-------------[繰り返し処理/h]---------------------------------



#--------------[meta/h]---------------------
category = input("業種を入力 : ")
myId = input("IDを入力 : ")
myPass = input("passを入力 : ")
input_index = input("取得する企業情報数 : ")
#--------------[meta/r]---------------------

#--------------[ドライバーの指定/h]---------------------
driver = webdriver.Chrome("/usr/local/bin/chromedriver")
driver.get("https://job.mynavi.jp/21/pc/login?func=PCtop")
#--------------[ドライバーの指定/r]---------------------

login(myId, myPass, input_index) #mynavi login

search(category) #業種search

# 企業ページオープン
factory_name = driver.find_element_by_id('corpNameLink[0]')
factory_name.click()

# ウィンドウハンドルの変更
handle_array = driver.window_handles
driver.switch_to.window(handle_array[1])

fac_info() #ループ処理

# タブクローズ
driver.close()
driver.switch_to.window(handle_array[0])
driver.close()